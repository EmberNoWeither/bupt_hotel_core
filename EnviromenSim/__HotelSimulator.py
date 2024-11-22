from .__base__.__BaseSimulator import BaseSimulator
from .__RoomSimulator import RoomSimulator as Room
from .__RoomerSim import RoomerSimulator as Roomer
from utils.BaseOperators import list_attr_Add
from utils.BaseValue import ROOM_STATUS,AC_CTROL,AC_MODE,AC_STATUS,PA,WIND_SPEED,base_ratio,is_simu
from utils.pymysql.database_operations import *
from InforTransfer.__ControlMsg import ControlMsg
import datetime
import time
import math
import copy
from functools import cmp_to_key
import pandas as pd
import threading
from openpyxl import load_workbook

WOKRKING_LOCK = threading.Lock()    # 控制和计算费用模拟等进行互斥以防万一


def control_sortcmp(cmd1:ControlMsg, cmd2:ControlMsg):
    """调度算法：
        被执行次数越少的先执行（防止顾客冻死或热死）
        执行次数相等则风速优先级
        风速优先级相同则时间优先级
    """
    
    if cmd1.working_times < cmd2.working_times:
        return -1
    elif cmd1.working_times > cmd2.working_times:
        return 1
    elif cmd1.working_times == cmd2.working_times:
        if cmd1._get_level() > cmd2._get_level():
            return -1
        elif cmd1._get_level() < cmd2._get_level():
            return 1
        elif cmd1._get_level() == cmd2._get_level():
            if cmd1._get_time() <= cmd2._get_time():
                return 1
            elif cmd1._get_time() > cmd2._get_time():
                return -1


class HotelSimulator(BaseSimulator):
    def __init__(self, init_rooms_num = 6, init_rooms_temp = [15 for i in range(6)]) -> None:
        super().__init__()
        self.__property = {
        'rooms_id' : [i+1 for i in range(init_rooms_num)],
        'out_tmp' : 15,
        'rooms_tmp' : {},
        'administ_num' : 1,
        'super_administ_num' : 1,
        'roomer_num' : 0,
        'roomers' : [],
        'rooms' : [],   
        }
        self.is_simu = False
        self.base_ratio = 1.0
        self.command_queue = []
        
        if len(init_rooms_temp) != init_rooms_num:
            assert "初始化房间温度列表长度不等于房间数目"
        
        init_rooms_temp = [11, 10, 15, 18, 12, 14]
        for i in range(init_rooms_num):
            self.__property['rooms_tmp'][i] = init_rooms_temp[i]
            self.__property['rooms'].append(Room(i, init_rooms_temp[i], self.__property['out_tmp']))
        
        # 根据数据库信息初始化
        roomers_data = load_guest()
        print(roomers_data)
        try:
            if roomers_data:
                for roomer_data in roomers_data:
                    roomer = Roomer(roomer_data['room_id'], roomer_data['phone_number'], roomer_data['id_card'], roomer_data['sign_time'],
                                    total_cost=roomer_data['total_cost'], power_usages=roomer_data['power_usage'], air_cost=roomer_data['air_cost'])
                    self.__property['roomer_num'] += 1
                    self.__property['roomers'].append(roomer)
                    for room in self.__property['rooms']:
                        if roomer(attr='room_id') == room(attr='room_id'):
                            room.check_in_room(roomer)
                            break
        except:
            pass
        
    def __call__(self, *args: any, **kwds: any) -> any:
        return super().__call__(*args, **kwds)
    
    
    def write_commandTable(self, outputs_list, command_line, idx=0):
        wb = load_workbook(r"D:\bupt_hotel_core\\outputs.xlsx")
        sheet = wb.get_sheet_by_name(r"制热测试用例")
        
        base_tuple = (4, 7)
        # outputs_list [{'room_id':xx, 'room_temp':xx, 'target_temp':xx, 'wind':xx,'now_cost':xx}]
        row = idx + base_tuple[0]
        for outputs in outputs_list:
            id = outputs['room_id']
            column = base_tuple[1] + (id-1)*4
            sheet.cell(row=row, column=column).value = outputs['room_temp']
            sheet.cell(row=row, column=column+1).value = outputs['target_temp']
            
            if outputs['wind'] == 'middle':
                sheet.cell(row=row, column=column+2).value = '中'
            elif outputs['wind'] == 'slow':
                sheet.cell(row=row, column=column+2).value = '低'
            elif outputs['wind'] == 'high':
                sheet.cell(row=row, column=column+2).value = '高'
                
            sheet.cell(row=row, column=column+3).value = outputs['now_cost']
            
        for i, command_id in enumerate(command_line):
            sheet.cell(row=row, column=28+i).value = command_id
        
        wb.save(r'D:\bupt_hotel_core\\outputs.xlsx')
        
    
    def stringcommand_read(self, table_command):
        
        if table_command == '开机':
                    command = {
                        'type': AC_CTROL.Open,
                        'value': {
                            'mode' : AC_MODE.Hot,
                            'set_tmp' : 22,
                            'wind_spd':WIND_SPEED.middle,
                            'wind_drc':None,
                        },
                        'authority' : PA.roomer
                    }
                    
        elif table_command == '关机':
            command = {
                'type': AC_CTROL.Close,
                'value': {
                    'mode' : AC_MODE.Hot,
                    'set_tmp' : None,
                    'wind_spd':WIND_SPEED.middle,
                    'wind_drc':None,
                },
                'authority' : PA.roomer
            }
        elif table_command == '高':
            command = {
                'type': AC_CTROL.SetWindSpd,
                'value': {
                    'mode' : AC_MODE.Hot,
                    'set_tmp' : None,
                    'wind_spd':WIND_SPEED.high,
                    'wind_drc':None,
                },
                'authority' : PA.roomer
            }
        elif table_command == '中':
            command = {
                'type': AC_CTROL.SetWindSpd,
                'value': {
                    'mode' : AC_MODE.Hot,
                    'set_tmp' : None,
                    'wind_spd':WIND_SPEED.middle,
                    'wind_drc':None,
                },
                'authority' : PA.roomer
            }
        elif table_command == '低':
            command = {
                'type': AC_CTROL.SetWindSpd,
                'value': {
                    'mode' : AC_MODE.Hot,
                    'set_tmp' : None,
                    'wind_spd':WIND_SPEED.slow,
                    'wind_drc':None,
                },
                'authority' : PA.roomer
            }
            
        return command
    
    def get_commandTable(self, line):
        file = pd.read_excel(r'D:\bupt_hotel_core\\验收用例.xlsx', sheet_name=1, usecols=[0,1,2,3,4,5])
        # file
        table_commands_set = ['开机', '关机', '高', '中', '低']
        commands = file.iloc[2 + line].fillna('N')

        room_commands = []
        # 房间 指令生成
        for idx, table_command in enumerate(commands[1:6]):
            if isinstance(table_command, int):
                command = {
                        'type': AC_CTROL.SetTemp,
                        'value': {
                            'mode' : AC_MODE.Hot,
                            'set_tmp' : table_command,
                            'wind_spd':None,
                            'wind_drc':None,
                        },
                        'authority' : PA.roomer
                    }
                
                self.Hotel_ControlProcess(idx+1, command)
                continue
            
            if table_command in table_commands_set and isinstance(table_command, str):
                command = self.stringcommand_read(table_command)
                self.Hotel_ControlProcess(idx+1, command)
            elif table_command != 'N':
                # 表中组合指令数字在前，风速在后
                command1 = int(table_command.split('，')[0])
                command2 = table_command.split('，')[1]
                command = self.stringcommand_read(command2)
                command['value']['set_tmp'] = command1
                self.Hotel_ControlProcess(idx+1, command)
    
    
    def Hotel_RealSimulating(self):
        while True:
            with WOKRKING_LOCK:
                self.Hotel_ControlCommandScheduleCirculating()
                
            self.Hotel_WatchingRoom([1,2,3,4,5])
            time.sleep(10)
            
    
    def Hotel_Simulating(self):
        command_id = 0
        while True:
            if self.is_simu:
                self.get_commandTable(command_id)
                if command_id == 26:
                    break
                
                
            with WOKRKING_LOCK:
                outputs_list, room_line = self.Hotel_ControlCommandScheduleCirculating()
                
            if self.is_simu:
                self.write_commandTable(outputs_list=outputs_list, command_line=room_line, idx=command_id)
                command_id += 1
                    
            self.Hotel_WatchingRoom([1,2,3,4,5])
            # print(command_line)

            time.sleep(10)
            
        self.Hotel_RealSimulating()            
            
    # 仅在关机指令后调用，用以生成详单
    def Hotel_ControlFinishCheck(self, room_id):
        for command in self.command_queue:
            # 如果出现有已经完成的服务的情况（理论上不会，以防代码错误）
            if command.finish:
                continue
            
            command_id = command._get_room_id()
            for room in self.__property['rooms']:
                if command_id == room(attr='room_id') and command_id == room_id:
                    # 检查温度是否已达标
                    command.finish = True   # 服务请求已完成
                    command.endTime = datetime.datetime.now()
                    specification = {
                        'room_id' : room_id,
                        'request_time' : command._get_time(),
                        'serve_time' : command.serve_time,
                        'serve_start' : command.serve_start,
                        'serve_end' : command.endTime,
                        'wait_time' : command.waiting_time,
                        'request_period' : float((command.endTime - command._get_time()).total_seconds()),
                        'wind_spd' : room.ac(attr='wind_spd').value,
                        'now_air_cost' : room.controller._get_cost(),
                        'now_cost_ratio' : self.base_ratio
                        }
                    
                    room.specifications_save(specification)
                    
                    # 去除该请求
                    self.command_queue.remove(command)
                    
                    
    def Hotel_WatchingRoom(self, id_list):
        print("************************************************************************************")
        for room_id in id_list:
            for room in self.__property['rooms']:
                if room_id == room(attr='room_id'):
                    temp = room(attr='room_tmp')
                    print("-----------------------------------------------------------------------------")
                    print("room_id: %d"%room_id + " room_temp: %f"%temp + " set_tmp: %f"%room.ac(attr='set_tmp')
                          +" status:"+room.ac(attr='status').name+" mode:"+room.ac(attr='mode').name+" wind:"+room.ac(attr='wind_spd').name
                          +" is_ava:"+room(attr='status').name)
        print("-----------------------------------------------------------------------------")
                    
                    
    def Hotel_GetRoomStatus(self, room_id):
        room_status = {}
        for room in self.__property['rooms']:
            if room_id == room(attr='room_id'):
                ac_is_on = True if room.ac(attr='status') != AC_STATUS.Stop else False
                ac_mode = ''
                if room.ac(attr='mode') == AC_MODE.Hot:
                    ac_mode = 'hot'
                elif room.ac(attr='mode') == AC_MODE.Cold:
                    ac_mode = 'cold'
                
                room_status = {
                    'is_on' : ac_is_on,
                    'is_ava' : False if room(attr='roomer') else True,
                    'mode' : ac_mode,
                    'cur_temp' : room(attr='room_tmp'),
                    'tar_temp' : room.ac(attr='set_tmp'),
                    'wind' : room.ac(attr='wind_spd').value - 1,
                    'now_cost' : room(attr='roomer')(attr='air_cost') if room(attr='roomer') else 0.0
                }

        return room_status
    

    
    def Hotel_ControlCommandScheduleCirculating(self):  # 该函数每分钟调用一次
        # 扫描请求队列
        if len(self.command_queue) == 0:
            for room in self.__property['rooms']:
                air_cost, power_usages, room_id, room_temp = room.tmp_simulating(1) # 执行指定间隔时间的温度模拟与计费
                self.__property['rooms_tmp'][room_id] = room_temp
            return [], []    # 无请求需调度
        
        # 优先级扫描排序——先按风速大小，再按时间先后
        self.command_queue = sorted(self.command_queue, key=cmp_to_key(control_sortcmp))    # 排序后头部为最高优先级
        
        room_line = [command._get_room_id() for command in self.command_queue]
        print(room_line)
        
        command_running = []
        outputs_list = []
        # 前3个为服务队列，后2个为等待队列
        i = 0
        for command in self.command_queue:
            command_running.append(command._get_room_id())
            outputs = {}
            if i < 3:
                command.update_serve_start()
                for room in self.__property['rooms']:
                    if room(attr='room_id') == command._get_room_id():
                        air_cost, power_usages, room_id, room_temp = room.tmp_simulating(self.base_ratio) # 执行指定间隔时间的温度模拟与计费
                        
                        outputs['room_id'] = room_id
                        outputs['room_temp'] = room_temp
                        outputs['target_temp'] = room.ac(attr='set_tmp')
                        outputs['wind'] = room.ac(attr='wind_spd').name
                        outputs['now_cost'] = air_cost
                        
                        #TODO 根据已服务时间和当前运行时间写表等
                        self.__property['rooms_tmp'][room_id] = room_temp
                        command.working_period -= 1
                        command.serve_time += 1
                        # 时间片用完下处理机，重置时间片
                        if command.working_period == 0:
                            command.working_period = 2
                            command.working_times += 1  # 增加运行次数
            else:
                
                # TODO 被扔进等待队列的指令同样也要模拟，要进行回温并生成一条详单
                for room in self.__property['rooms']:
                    if room(attr='room_id') == command._get_room_id():
                        room.ac.change_status(AC_STATUS.Waiting)
                        air_cost, power_usages, room_id, room_temp = room.tmp_simulating(1) # 执行指定间隔时间的温度模拟与计费
                        self.__property['rooms_tmp'][room_id] = room_temp
                        room.ac.change_status(AC_STATUS.Running)    # 执行完后将空调变回原状态（只要没关机就先认为在运行）
                        
                        outputs['room_id'] = room_id
                        outputs['room_temp'] = room_temp
                        outputs['target_temp'] = room.ac(attr='set_tmp')
                        outputs['wind'] = room.ac(attr='wind_spd').name
                        outputs['now_cost'] = air_cost
                        
                        if command.is_working == True:  # 第一次进入等待队列，进行详单生成(认为一段服务结束)
                            # command.finish = True   # 服务请求已完成
                            command.endTime = datetime.datetime.now()
                            specification = {
                                'room_id' : room_id,
                                'request_time' : command._get_time(),
                                'serve_time' : command.serve_time,
                                'serve_start' : command.serve_start,
                                'serve_end' : command.endTime,
                                'wait_time' : command.waiting_time,
                                'request_period' : float((command.endTime - command._get_time()).total_seconds()),
                                'wind_spd' : room.ac(attr='wind_spd').value,
                                'now_air_cost' : room.controller._get_cost(),
                                'now_cost_ratio' : self.base_ratio
                            }
                            room.specifications_save(specification)
                
                command.sleep()
                # 进等待队列的房间，时间片重置
                command.working_period = 2
                command.waiting_time += 1
            outputs_list.append(outputs)
            i += 1
            
            
        for room in self.__property['rooms']:
            if room(attr='room_id') not in command_running: # 对于无指令房间同样要进行模拟
                air_cost, power_usages, room_id, room_temp = room.tmp_simulating(1) # 执行指定间隔时间的温度模拟与计费
                self.__property['rooms_tmp'][room_id] = room_temp
                
        
        return outputs_list, room_line
            
    
    def Hotel_ControlProcess(self, room_id, command):
        with WOKRKING_LOCK:
            for room in self.__property['rooms']:
                if room_id == room(attr='room_id'):
                    # 如果是关空调指令，进行服务完成检查，并记录
                    if command['type'] == AC_CTROL.Close:
                        command['value']['set_tmp'] = 22
                        self.Hotel_ControlFinishCheck(room_id)
                        room.roomer_control(command, datetime.datetime.now())
                        break
                    
                    # 先检查等待队列中是否有原本的控制服务请求，有则进行修改，无则添加
                    last = None
                    if len(self.command_queue) != 0:
                        for last_command in self.command_queue:
                            if last_command._get_room_id() == room_id:
                                if last_command.finish: # 若上个请求已经完成 即关过机 --理论上该请求应该被移除，但万一……
                                    pass
                                else:
                                    last = copy.deepcopy(last_command)
                                    
                                self.command_queue.remove(last_command)
                
                    control_msg = room.roomer_control(command, datetime.datetime.now())
                    
                    
                    # 请求未完成，更新请求内容,保留调度有关信息 -> 业务逻辑更改为若存在先前执行中指令，进行之前的详单生成
                    if last:
                        last.finish = True   # 服务请求已完成
                        last.endTime = datetime.datetime.now()
                        specification = {
                            'room_id' : room_id,
                            'request_time' : last._get_time(),
                            'serve_time' : last.serve_time,
                            'serve_start' : last.serve_start,
                            'serve_end' : last.endTime,
                            'wait_time' : last.waiting_time,
                            'request_period' : float((last.endTime - last._get_time()).total_seconds()),
                            'wind_spd' : room.ac(attr='wind_spd').value,
                            'now_air_cost' : room.controller._get_cost(),
                            'now_cost_ratio' : self.base_ratio
                        }
                        room.specifications_save(specification)
                        
                        control_msg.update_msg(last)

                    self.command_queue.append(control_msg)
                    break
                
                
    def Hotel_GetSpecifications(self, room_id, id_card):
        return get_specifications(room_id, id_card)
    
    
    def Check_In(self, rooms_id, phone_number, id_card):
        # 数据库查询空房
        room_data, empty_nums = empty()
        if empty_nums == 0:
            v = "很抱歉，酒店已住满，暂无房间可用！"
            print(v)
            return v
        
        if str(rooms_id) not in room_data['ava_room']:
            # 查询下对应房客信息是否存在
            roomers_data = load_guest()
            exist = False
            if roomers_data:
                for room_data in roomers_data:
                    if room_data['room_id'] == rooms_id:
                        exist = True
                        break
                    
                if exist:
                    v = "该房间刚刚已被占用，请重新选择房间！"
                    print(v)
                    return v
            
        roomer = Roomer(rooms_id, phone_number, id_card, datetime.datetime.now())
        self.__property['roomer_num'] += 1
        self.__property['roomers'].append(roomer)
        
        for room in self.__property['rooms']:
            if roomer(attr='room_id') == room(attr='room_id'):
                room.check_in_room(roomer)
                break
        
        return 1
    
    
    def Check_Out(self, rooms_id, phone_number, id_card):
        check_out_database, bill = data_check_out(id_card)
        if check_out_database:
            # 尝试查找系统缓存是否有该记录，进行删除
            for roomer in self.__property['roomers']:
                if rooms_id == roomer(attr='room_id'):
                    for room in self.__property['rooms']:
                        if roomer(attr='room_id') == room(attr='room_id'):
                            room.check_out_room()
                            room.controller.checkout_costs()
                            self.__property['roomers'].remove(roomer)
                            self.__property['roomer_num'] -= 1
                            break
                    
        if check_out_database:
            return bill
        else:
            print("所要退房的房客不存在！")
            return bill
                    